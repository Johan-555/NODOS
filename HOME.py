from openpyxl import load_workbook
from collections import Counter
import os

# Cargar el archivo Excel
archivo = "AUTOMATISMO-NODO.xlsx"
workbook = load_workbook(filename=archivo)
hoja = workbook.active  # Seleccionar la hoja activa

conteo_valores = {}

# Contar los valores en la columna 3, comenzando desde la fila 2
for fila in hoja.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True):
    for valor in fila:
        if valor is not None:  # Ignorar celdas vacías
            if valor in conteo_valores:
                conteo_valores[valor] += 1
            else:
                conteo_valores[valor] = 1

# Identificar y mostrar los valores repetidos
repetidos = {k: v for k, v in conteo_valores.items() if v > 1}

# Escribir el encabezado para los nodos repetidos en la celda D1
hoja.cell(row=1, column=4).value = "Nodos Repetidos"
hoja.cell(row=1, column=5).value = "Cantidad de Repeticiones"
hoja.cell(row=1, column=6).value = "Nodos con mayor repetición"
hoja.cell(row=1, column=7).value = "Nodos que más se repiten en su respectivo mes"
hoja.cell(row=1, column=8).value = "Total de repeticiones por mes"
hoja.cell(row=1, column=9).value = "Top 10 nodos por mes" 
hoja.cell(row=1, column=11).value = "Fecha cargue de Nodos" 

# Escribir los valores repetidos en la columna 4, comenzando desde la fila 2
row = 2
for valor, cuenta in repetidos.items():
    hoja.cell(row=row, column=4).value = valor
    hoja.cell(row=row, column=5).value = cuenta
    row += 1

# Función para detectar y ordenar las combinaciones de las columnas 4 y 5
def detectar_combinaciones_ordenadas():
    combinaciones = []
    row = 2
    while hoja.cell(row=row, column=4).value is not None:
        nodo = hoja.cell(row=row, column=4).value
        repeticiones = hoja.cell(row=row, column=5).value
        if nodo is not None and repeticiones is not None:
            combinaciones.append((nodo, repeticiones))
        row += 1
    combinaciones_ordenadas = sorted(combinaciones, key=lambda x: x[1], reverse=True)
    row = 2
    for nodo, repeticiones in combinaciones_ordenadas:
        hoja.cell(row=row, column=6).value = f"{nodo} - {repeticiones}"
        row += 1

# Función para contar las combinaciones de columnas 2 y 3 y ordenarlas por mayor repetición
def contar_y_ordenar_combinaciones_meses_nodos():
    combinaciones_counter = Counter()
    row = 2
    while hoja.cell(row=row, column=2).value is not None and hoja.cell(row=row, column=3).value is not None:
        mes = hoja.cell(row=row, column=2).value
        nodo = hoja.cell(row=row, column=3).value
        if mes is not None and nodo is not None:
            combinaciones_counter[(mes, nodo)] += 1
        row += 1
    combinaciones_ordenadas = combinaciones_counter.most_common()
    row = 2
    for (mes, nodo), repeticiones in combinaciones_ordenadas:
        hoja.cell(row=row, column=7).value = f"{mes} - {nodo}: {repeticiones} veces"
        row += 1

# Función para sumar las repeticiones totales de cada mes y ordenarlas
def sumar_repeticiones_por_mes():
    suma_por_mes = Counter()
    row = 2
    while hoja.cell(row=row, column=7).value is not None:
        celda_valor = hoja.cell(row=row, column=7).value
        if celda_valor:
            mes, resto = celda_valor.split(" - ", 1)
            repeticiones_texto = resto.split(": ")[1].replace(" veces", "")
            repeticiones = int(repeticiones_texto)
            suma_por_mes[mes] += repeticiones
        row += 1
    suma_ordenada = suma_por_mes.most_common() 
    row = 2
    for mes, total_repeticiones in suma_ordenada:
        hoja.cell(row=row, column=8).value = f"{mes}: {total_repeticiones} repeticiones"
        row += 1

# Nueva función para generar el top 10 de los nodos más repetidos por mes
def top_10_nodos_por_mes():
    nodos_por_mes = {}

    # Leer las combinaciones de la columna 7
    row = 2
    while hoja.cell(row=row, column=7).value is not None:
        celda_valor = hoja.cell(row=row, column=7).value
        if celda_valor:
            mes, resto = celda_valor.split(" - ", 1)
            nodo, repeticiones_texto = resto.split(": ")
            repeticiones = int(repeticiones_texto.replace(" veces", ""))
            if mes not in nodos_por_mes:
                nodos_por_mes[mes] = []
            nodos_por_mes[mes].append((nodo, repeticiones))
        row += 1

    # Ordenar nodos por mes y seleccionar el top 10
    row = 2
    for mes, nodos in nodos_por_mes.items():
        top_10 = sorted(nodos, key=lambda x: x[1], reverse=True)[:10]
        hoja.cell(row=row, column=9).value = f"Top 10 de {mes}:"
        row += 1
        for nodo, repeticiones in top_10:
            hoja.cell(row=row, column=9).value = f"{nodo} - {repeticiones} veces"
            row += 1

# Función para emparejar datos de columna 3 con los de columna 10 y escribir en columna 11
def emparejar_y_ordenar_col3_y_col10():
    datos = []
    row = 2
    while hoja.cell(row=row, column=3).value is not None or hoja.cell(row=row, column=10).value is not None:
        valor_col3 = hoja.cell(row=row, column=3).value
        valor_col10 = hoja.cell(row=row, column=10).value
        if valor_col3 is not None and valor_col10 is not None:
            datos.append(f"{valor_col3} - {valor_col10}")
        elif valor_col3 is not None:
            datos.append(valor_col3)
        elif valor_col10 is not None:
            datos.append(valor_col10)
        row += 1

    # Ordenar los datos en función de la fecha en la parte final del string, en orden descendente
    datos.sort(key=lambda x: x.split(" - ")[-1], reverse=True)

    # Escribir los datos ordenados en la columna 11
    for idx, valor in enumerate(datos, start=2):
        hoja.cell(row=idx, column=11).value = valor

# Ejecutar las funciones
detectar_combinaciones_ordenadas()
contar_y_ordenar_combinaciones_meses_nodos()
sumar_repeticiones_por_mes()
top_10_nodos_por_mes()
emparejar_y_ordenar_col3_y_col10()

# Guardar el archivo modificado
updated_file_path = "NODOS_modificado.xlsx"
workbook.save(updated_file_path)

# Abrir el archivo Power BI    
power_bi_file = "NODOS UPDATE.pbix"
os.startfile(power_bi_file)

# Instrucciones para el usuario
print("\nINSTRUCCIONES:")
print("1. En Power BI, selecciona 'Actualizar' para recargar los datos.")
print("2. Usa las gráficas existentes para analizar los datos del nuevo archivo.")
print("3. Guarda el archivo de Power BI si es necesario.") 
